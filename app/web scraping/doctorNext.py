import requests
from openpyxl import load_workbook
import math

class DoctorNext:
    def __init__(self,url) -> None:
        self.url = url
    
    def getResponse(self, payload, page):
        payload['page'] = page
        response = requests.post(url=self.url, json=payload)
        return response.json()
    
    def getPages(self, payload):
        response_json = self.getResponse(payload,1)
        return math.ceil(response_json['meta']['totalRecords']/response_json['meta']['pageSize'])    
    
    def getDoctorsData(self, payload):
        pages = self.getPages(payload)
        doctorData = []
        for page in range(1, pages+1):
            
            response_json = self.getResponse(payload,page)
            doctors = response_json['data']
            
            for dr in doctors:
                data = [
                    dr.get('persianFirstName', ""),
                    dr.get('persianLastName', ""),
                    dr.get('nezamCode', ""),
                    dr.get('gender', ""),
                    dr['speciality'].get('name', "") if 'speciality' in dr else "",
                    dr.get('rate', "")
                ]
                doctorData.append(data)
            print(len(doctorData))
        return doctorData
    
    def getOfficesData(self, payload):
        pages = self.getPages(payload)
        officeData = []
        for page in range(1, pages+1):
            
            response_json = self.getResponse(payload,page)
            doctors = response_json['data']

            for dr in doctors:
                for ofc in dr.get('offices', []):
                    element = []
                    element.append(dr.get('persianFirstName', ""))
                    element.append(dr.get('persianLastName', ""))
                    element.append(dr.get('nezamCode', ""))
                    element.append(ofc.get('name', ""))
                    if ofc['city'] is not None: 
                        if ofc['city']['state'] is not None:                 
                            element.append(ofc['city']['state'].get('name', ""))
                        else:
                            element.append("")
                        element.append(ofc['city'].get('name', ""))
                    else:
                        element.append("")    
                    element.append(ofc.get('address', ""))
                    element.append(ofc.get('latitude', ""))
                    element.append(ofc.get('longitude', ""))
                    element.append(ofc.get('telephone', ""))
                officeData.append(element)
            return officeData

    def writeToExcel(self,excel_file, sheet, data):
        wb = load_workbook(excel_file)
        ws = wb[sheet]    
        row = ws.max_row 
        for row_data in data:
            for col, value in enumerate(row_data, start=1):
                ws.cell(row=row, column=col, value=value)
            row += 1
        wb.save(excel_file)        
        

if __name__=="__main__": 
    
    url = 'https://cyclops.drnext.ir/v1/website/search'
    payload = {
        "sort": {
            "name": "sortFreeTimeValue",
            "order": "ASC"
        },
        "query": "",
        "page": 1,
        "speciality": "women-oncology-fellowship",
        "officeType": [],
        "gender": []
    }
    file_name = "test.xlsx"
    dr = DoctorNext(url)
    drData = dr.getDoctorsData(payload)
    print(dr.getPages(payload))
    dr.writeToExcel(file_name,"Sheet1",drData)